#!/usr/bin/env python
"""xlsx → 중간 CSV. 코어 자료 반입의 앞단이다 (P17 4절).

**왜 두 단인가.** 뒷단(`ops/import_coredata.py`)은 규약대로 컨테이너 안에서
도는데, 거기서 xlsx 를 바로 읽으려면 웹 이미지에 `openpyxl` 이 들어가야 하고
NAS 공유(`/nfs/temp-share`)까지 컨테이너가 봐야 한다. 반입 하나 때문에 늘리기엔
넓다. 그리고 **1.1 의 단위 판단은 눈에 보이는 파일로 남아야 한다** — 그것이
`coredata/mapping.toml` 이고, 이 스크립트는 그 표만 따른다.

**Django 를 안 부르고 DB 도 안 만진다** — `ops/export_review.py`·
`ops/backup_db.py` 와 같은 자리라 호스트 venv 로 돈다.

    python tools/coredata_extract.py \
        --xlsx-dir /nfs/temp-share/DiaRUGA/coredata \
        --out      /data3/DiaRUGA/coredata

내는 것은 코어마다 둘이다:

    <지역>-<지점>.series.csv   key,label,unit,default_on,sort_order,origin
    <지역>-<지점>.points.csv   key,depth_mm,value

**깊이는 mm 정수로 낸다.** DB 가 그렇게 든다 — `(항목, 깊이)` 가 유일 제약의
열쇠라서 부동소수면 안 된다 (`CorePoint` 머리말).
"""
import argparse
import csv
import sys
import tomllib
from pathlib import Path

# **임포트에서 멈추지 않는다.** 시험이 `read_block`·`_to_mm` 을 xlsx 없이
# 부르는데(엑셀은 `iter_rows` 하나로만 쓴다), 여기서 `sys.exit` 하면 그 시험이
# openpyxl 을 깔아야만 도는 것이 된다 — 이 파일은 호스트 전용이고 웹 이미지에는
# 안 들어간다. 없으면 실제로 xlsx 를 열 때 말한다.
try:
    import openpyxl
except ImportError:                                          # pragma: no cover
    openpyxl = None

# 매핑표의 `expect_max_cm` 에서 이만큼 벗어나면 멈춘다. **단위를 잘못 읽으면
# 열 배로 어긋나므로** 넉넉히 잡아도 그 사고는 잡힌다. 코어 길이 자체가
# 어림값이라 좁게 잡으면 멀쩡한 반입이 멈춘다.
DEPTH_LO, DEPTH_HI = 0.5, 2.0


def _num(v):
    """숫자만 통과시킨다. `TYPE`·`Munsell C Hue` 같은 글자 칸은 여기서 걸린다."""
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def _to_mm(depth: float, unit: str) -> int:
    """깊이를 mm 정수로. **반올림한다** — 0.1 cm 가 1 mm 다."""
    return round(depth * 10) if unit == "cm" else round(depth)


def read_block(ws, blk: dict) -> dict[str, dict[int, float]]:
    """블록 하나에서 `{key: {depth_mm: value}}` 를 뽑는다.

    **열은 번호로 짚는다.** `MS` 의 `Point`/`Whole`, `Spectro` 의 `SCE`/`SCI` 가
    같은 머리글을 반복해서 이름으로 짚으면 뒤엣것이 앞엣것을 덮는다.
    """
    dcol, dunit = blk["depth_col"], blk["depth_unit"]
    cols = [(int(c[0]), c[1]) for c in blk["columns"]]
    out: dict[str, dict[int, float]] = {k: {} for _, k in cols}
    same, clash = 0, []
    for row in ws.iter_rows(min_row=blk["header_row"] + 1, values_only=True):
        depth = _num(row[dcol - 1]) if len(row) >= dcol else None
        if depth is None:
            continue
        mm = _to_mm(depth, dunit)
        for col, key in cols:
            val = _num(row[col - 1]) if len(row) >= col else None
            # **값이 없는 깊이는 안 넣는다.** 넣으면 화면이 그 구간을 이어
            # 그려 안 잰 구간이 잰 것처럼 뜬다.
            if val is None:
                continue
            if mm in out[key]:
                # **같은 값이면 붙여넣기가 겹친 것이고, 다르면 사람이 정할
                # 일이다.** `RS14-GC04` 의 `MS` 가 221~260 cm 40점을 한 번 더
                # 들고 있는데 값이 글자 그대로 같다 — 그건 버려도 된다. 값이
                # 다르면 어느 쪽이 맞는지 이 스크립트가 고를 수 없다.
                if out[key][mm] != val:
                    clash.append((key, mm, out[key][mm], val))
                else:
                    same += 1
                continue
            out[key][mm] = val
    if same:
        print(f"      같은 깊이·같은 값이 {same}개 겹쳐 있어 한 번만 넣습니다 "
              f"(시트 {blk['sheet']})")
    if clash:
        print(f"      !! 같은 깊이에 다른 값이 있습니다 (시트 {blk['sheet']}, "
              f"{len(clash)}건) — 어느 쪽이 맞는지 사람이 정해야 합니다")
        for key, mm, a, b in clash[:5]:
            print(f"         {key} {mm / 10:g} cm: {a:g} vs {b:g}")
        raise ValueError(f"{blk['sheet']}: 같은 깊이에 다른 값")
    return out


def extract(core: dict, xlsx_dir: Path, out_dir: Path) -> int:
    name = f"{core['site']}-{core['locality']}"
    src = xlsx_dir / core["file"]
    if not src.exists():
        print(f"  !! 원본이 없습니다: {src}")
        return 1
    if openpyxl is None:
        print("  !! openpyxl 이 없습니다 — pip install openpyxl")
        return 1
    print(f"  {name}  ← {core['file']}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    default_on = set(core.get("default_on", []))
    meta: dict[str, dict] = {}
    points: dict[str, dict[int, float]] = {}
    order = 0
    for blk in core["block"]:
        if blk["sheet"] not in wb.sheetnames:
            print(f"      !! 시트가 없습니다: {blk['sheet']}")
            return 1
        try:
            got = read_block(wb[blk["sheet"]], blk)
        except ValueError:
            # `read_block` 이 이미 무엇이 어긋났는지 적었다. 여기서 스택을
            # 쏟으면 그 줄이 묻힌다.
            wb.close()
            return 1
        for i, col in enumerate(blk["columns"]):
            key, label, unit = col[1], col[2], col[3]
            if key in meta:
                print(f"      !! key 가 겹칩니다: {key}")
                return 1
            order += 1
            meta[key] = {
                "key": key, "label": label, "unit": unit,
                "default_on": 1 if key in default_on else 0,
                "sort_order": blk.get("sort_order", 100) + i,
                "origin": f"{core['file']}::{blk['sheet']}",
            }
            points[key] = got[key]
    wb.close()

    # **사람이 적어 둔 코어 길이로 자기를 검사한다.** 단위를 잘못 읽으면 열
    # 배로 어긋나므로 여기서 걸린다 — 매핑표가 짐작을 막는 것과 짝이다.
    expect = core.get("expect_max_cm")
    bad = []
    if expect:
        for key, pts in points.items():
            if not pts:
                continue
            deep = max(pts) / 10
            if not (expect * DEPTH_LO <= deep <= expect * DEPTH_HI):
                bad.append((key, deep))
    if bad:
        print(f"      !! 코어 길이가 {expect} cm 인데 깊이가 벗어납니다 — "
              f"매핑표의 depth_unit 을 보세요")
        for key, deep in bad:
            print(f"         {key}: 최대 {deep:g} cm")
        return 1

    # 켜라고 적어 둔 항목이 실제로 있는가. 오타면 화면이 아무것도 안 켠 채
    # 뜨는데, 그것은 예외도 경고도 없는 종류의 고장이다.
    for key in sorted(default_on - set(meta)):
        print(f"      !! default_on 에 없는 항목이 적혀 있습니다: {key}")
        return 1

    out_dir.mkdir(parents=True, exist_ok=True)
    with (out_dir / f"{name}.series.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, ["key", "label", "unit", "default_on",
                               "sort_order", "origin"])
        w.writeheader()
        for key in sorted(meta, key=lambda k: meta[k]["sort_order"]):
            w.writerow(meta[key])
    n = 0
    with (out_dir / f"{name}.points.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["key", "depth_mm", "value"])
        for key in sorted(points, key=lambda k: meta[k]["sort_order"]):
            for mm in sorted(points[key]):
                w.writerow([key, mm, repr(points[key][mm])])
                n += 1
    empty = [k for k, v in points.items() if not v]
    print(f"      항목 {len(meta)}개 · 점 {n:,}개 → {out_dir / (name + '.*.csv')}")
    if empty:
        print(f"      점이 하나도 없는 항목 {len(empty)}개: {', '.join(sorted(empty))}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="코어 자료 xlsx → 중간 CSV")
    ap.add_argument("--mapping", default=str(Path(__file__).resolve().parent.parent
                                             / "coredata" / "mapping.toml"))
    ap.add_argument("--xlsx-dir", required=True)
    ap.add_argument("--out", required=True,
                    help="컨테이너도 보는 자리여야 한다 (/data3/DiaRUGA/coredata)")
    ap.add_argument("--only", default="", help="지역-지점 하나만 (RS14-GC04)")
    a = ap.parse_args()

    with open(a.mapping, "rb") as f:
        conf = tomllib.load(f)
    print(f"매핑표: {a.mapping}")
    rc = 0
    for core in conf["core"]:
        name = f"{core['site']}-{core['locality']}"
        if a.only and a.only != name:
            continue
        rc |= extract(core, Path(a.xlsx_dir), Path(a.out))
    return rc


if __name__ == "__main__":
    sys.exit(main())
